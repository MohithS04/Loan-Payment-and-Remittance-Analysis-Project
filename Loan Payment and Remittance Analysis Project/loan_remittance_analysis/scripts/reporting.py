"""
reporting.py — Automated Excel Report Builder

Generates three multi-sheet Excel reports with conditional formatting,
auto-filters, charts, and a reusable pivot template.
"""

import os
import sys

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# ── Style constants ─────────────────────────────────────────────────
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, num_cols: int):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_column_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    """Write a DataFrame to a worksheet starting at start_row."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER


# =========================================================================
#  Report 1 — Monthly Reconciliation Report
# =========================================================================

def generate_reconciliation_report(
    reconciliation_df: pd.DataFrame,
    discrepancy_df: pd.DataFrame,
    summary: dict,
    output_path: str | None = None,
) -> str:
    """Generate reconciliation_report.xlsx with 4 sheets."""
    if output_path is None:
        output_path = os.path.join(config.OUTPUT_REPORTS, "reconciliation_report.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    kpis = [
        ("KPI", "Value"),
        ("Total Payments Processed", summary["total_payments_processed"]),
        ("Matched", f"{summary['matched']} ({summary['matched'] / max(summary['total_payments_processed'], 1) * 100:.1f}%)"),
        ("Unmatched", summary["unmatched"]),
        ("Partial", summary["partial"]),
        ("Reversed", summary["reversed"]),
        ("Total Dollar Variance", f"${summary['total_dollar_variance']:,.2f}"),
        ("Discrepancy Rate", f"{summary['discrepancy_rate_pct']:.1f}%"),
        ("Errors Detected", len(discrepancy_df)),
    ]
    for r_idx, (k, v) in enumerate(kpis, 1):
        ws_summary.cell(row=r_idx, column=1, value=k).font = BOLD_FONT
        ws_summary.cell(row=r_idx, column=2, value=str(v)).font = NORMAL_FONT
        ws_summary.cell(row=r_idx, column=1).border = THIN_BORDER
        ws_summary.cell(row=r_idx, column=2).border = THIN_BORDER

    # Color-code KPIs
    ws_summary.cell(row=2, column=2).fill = GREEN_FILL  # Total
    ws_summary.cell(row=3, column=2).fill = GREEN_FILL  # Matched
    ws_summary.cell(row=4, column=2).fill = RED_FILL    # Unmatched
    ws_summary.cell(row=5, column=2).fill = YELLOW_FILL # Partial
    _auto_column_width(ws_summary)

    # ── Sheet 2: Discrepancies ──
    ws_disc = wb.create_sheet("Discrepancies")
    if len(discrepancy_df) > 0:
        _write_df_to_sheet(ws_disc, discrepancy_df)
        # Color-code severity
        if "severity" in discrepancy_df.columns:
            sev_col = list(discrepancy_df.columns).index("severity") + 1
            for r_idx in range(2, len(discrepancy_df) + 2):
                cell = ws_disc.cell(row=r_idx, column=sev_col)
                if cell.value == "High":
                    cell.fill = RED_FILL
                elif cell.value == "Medium":
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = GREEN_FILL
        _auto_column_width(ws_disc)
        ws_disc.auto_filter.ref = ws_disc.dimensions
    else:
        ws_disc.cell(row=1, column=1, value="No discrepancies found").font = BOLD_FONT

    # ── Sheet 3: By Pool ──
    ws_pool = wb.create_sheet("By Pool")
    pool_data = summary.get("by_pool", {})
    if pool_data:
        pool_rows = []
        for pool, stats in sorted(pool_data.items()):
            pool_rows.append({
                "Pool ID": pool,
                "Total": stats["total"],
                "Matched": stats["matched"],
                "Unmatched": stats["unmatched"],
                "Variance $": stats["variance"],
            })
        pool_df = pd.DataFrame(pool_rows)
        _write_df_to_sheet(ws_pool, pool_df)
        _auto_column_width(ws_pool)
    else:
        ws_pool.cell(row=1, column=1, value="No pool data available").font = BOLD_FONT

    # ── Sheet 4: By Trust Account ──
    ws_trust = wb.create_sheet("By Trust Account")
    trust_data = summary.get("by_trust_account", {})
    if trust_data:
        trust_rows = []
        for trust, stats in sorted(trust_data.items()):
            trust_rows.append({
                "Trust Account": trust,
                "Total": stats["total"],
                "Matched": stats["matched"],
                "Unmatched": stats["unmatched"],
                "Variance $": stats["variance"],
            })
        trust_df = pd.DataFrame(trust_rows)
        _write_df_to_sheet(ws_trust, trust_df)
        _auto_column_width(ws_trust)
    else:
        ws_trust.cell(row=1, column=1, value="No trust account data available").font = BOLD_FONT

    wb.save(output_path)
    return output_path


# =========================================================================
#  Report 2 — Collateral Compliance Report
# =========================================================================

def generate_collateral_report(
    variance_report: pd.DataFrame,
    compliance_summary: dict,
    output_path: str | None = None,
) -> str:
    """Generate collateral_report.xlsx with 3 sheets."""
    if output_path is None:
        output_path = os.path.join(config.OUTPUT_REPORTS, "collateral_report.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Compliance Summary ──
    ws_summary = wb.active
    ws_summary.title = "Compliance Summary"

    row = 1
    ws_summary.cell(row=row, column=1, value="Collateral Type").font = HEADER_FONT
    ws_summary.cell(row=row, column=1).fill = HEADER_FILL
    ws_summary.cell(row=row, column=1).border = THIN_BORDER

    col = 2
    status_names = set()
    for ctype_data in compliance_summary.get("by_collateral_type", {}).values():
        status_names.update(ctype_data.keys())
    status_names = sorted(status_names)

    for s in status_names:
        ws_summary.cell(row=row, column=col, value=s).font = HEADER_FONT
        ws_summary.cell(row=row, column=col).fill = HEADER_FILL
        ws_summary.cell(row=row, column=col).border = THIN_BORDER
        col += 1

    row = 2
    for ctype, stats in sorted(compliance_summary.get("by_collateral_type", {}).items()):
        ws_summary.cell(row=row, column=1, value=ctype).font = BOLD_FONT
        ws_summary.cell(row=row, column=1).border = THIN_BORDER
        col = 2
        for s in status_names:
            val = stats.get(s, 0)
            cell = ws_summary.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if s == "COMPLIANT":
                cell.fill = GREEN_FILL
            elif s == "NON-COMPLIANT":
                cell.fill = RED_FILL
            else:
                cell.fill = YELLOW_FILL
            col += 1
        row += 1

    # Totals row
    totals = compliance_summary.get("totals", {})
    ws_summary.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
    ws_summary.cell(row=row, column=1).fill = BLUE_FILL
    ws_summary.cell(row=row, column=1).border = THIN_BORDER
    col = 2
    for s in status_names:
        cell = ws_summary.cell(row=row, column=col, value=totals.get(s, 0))
        cell.font = BOLD_FONT
        cell.fill = BLUE_FILL
        cell.border = THIN_BORDER
        col += 1

    _auto_column_width(ws_summary)

    # ── Sheet 2: Detail ──
    ws_detail = wb.create_sheet("Detail")
    _write_df_to_sheet(ws_detail, variance_report)
    # Color-code compliance_status
    if "compliance_status" in variance_report.columns:
        cs_col = list(variance_report.columns).index("compliance_status") + 1
        for r_idx in range(2, len(variance_report) + 2):
            cell = ws_detail.cell(row=r_idx, column=cs_col)
            if cell.value == "COMPLIANT":
                cell.fill = GREEN_FILL
            elif cell.value == "NON-COMPLIANT":
                cell.fill = RED_FILL
            elif cell.value == "REQUIRES_REVIEW":
                cell.fill = YELLOW_FILL
    _auto_column_width(ws_detail)
    ws_detail.auto_filter.ref = ws_detail.dimensions

    # ── Sheet 3: Chart (Top 20 Outliers) ──
    ws_chart = wb.create_sheet("Chart")
    top_outliers = variance_report.nlargest(20, "variance_pct")[["security_id", "variance_pct"]].reset_index(drop=True)
    _write_df_to_sheet(ws_chart, top_outliers)
    _auto_column_width(ws_chart)

    if len(top_outliers) > 0:
        chart = BarChart()
        chart.title = "Top 20 Collateral Variance Outliers"
        chart.x_axis.title = "Security ID"
        chart.y_axis.title = "Variance %"
        chart.style = 10
        chart.width = 25
        chart.height = 15

        data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(top_outliers) + 1)
        cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(top_outliers) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_chart.add_chart(chart, "D2")

    wb.save(output_path)
    return output_path


# =========================================================================
#  Report 3 — Error Log Export
# =========================================================================

def generate_error_log_report(
    master_error_df: pd.DataFrame,
    accuracy: dict,
    output_path: str | None = None,
) -> str:
    """Generate error_log_report.xlsx with 2 sheets."""
    if output_path is None:
        output_path = os.path.join(config.OUTPUT_REPORTS, "error_log_report.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Master Error Log ──
    ws_master = wb.active
    ws_master.title = "Master Error Log"

    if len(master_error_df) > 0:
        sorted_df = master_error_df.sort_values("error_type").reset_index(drop=True)
        _write_df_to_sheet(ws_master, sorted_df)

        # Color-code severity
        if "severity" in sorted_df.columns:
            sev_col = list(sorted_df.columns).index("severity") + 1
            for r_idx in range(2, len(sorted_df) + 2):
                cell = ws_master.cell(row=r_idx, column=sev_col)
                if cell.value == "High":
                    cell.fill = RED_FILL
                elif cell.value == "Medium":
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = GREEN_FILL

        _auto_column_width(ws_master)
        ws_master.auto_filter.ref = ws_master.dimensions
    else:
        ws_master.cell(row=1, column=1, value="No errors detected").font = BOLD_FONT

    # ── Sheet 2: Error Trend / Accuracy Summary ──
    ws_trend = wb.create_sheet("Error Summary")
    summary_data = [
        ("Metric", "Value"),
        ("Total Ground Truth Errors", accuracy.get("total_ground_truth", 0)),
        ("Total Detected Errors", accuracy.get("total_detected", 0)),
        ("True Positives", accuracy.get("true_positives", 0)),
        ("False Positives", accuracy.get("false_positives", 0)),
        ("False Negatives", accuracy.get("false_negatives", 0)),
        ("Precision", f"{accuracy.get('precision', 0):.1f}%"),
        ("Recall", f"{accuracy.get('recall', 0):.1f}%"),
        ("F1 Score", f"{accuracy.get('f1_score', 0):.1f}%"),
    ]
    for r_idx, (k, v) in enumerate(summary_data, 1):
        ws_trend.cell(row=r_idx, column=1, value=k).font = BOLD_FONT if r_idx == 1 else NORMAL_FONT
        ws_trend.cell(row=r_idx, column=2, value=str(v)).font = NORMAL_FONT
        ws_trend.cell(row=r_idx, column=1).border = THIN_BORDER
        ws_trend.cell(row=r_idx, column=2).border = THIN_BORDER
        if r_idx == 1:
            ws_trend.cell(row=r_idx, column=1).fill = HEADER_FILL
            ws_trend.cell(row=r_idx, column=1).font = HEADER_FONT
            ws_trend.cell(row=r_idx, column=2).fill = HEADER_FILL
            ws_trend.cell(row=r_idx, column=2).font = HEADER_FONT

    # Error count by type
    if len(master_error_df) > 0:
        error_counts = master_error_df["error_type"].value_counts().reset_index()
        error_counts.columns = ["Error Type", "Count"]

        start_row = len(summary_data) + 3
        ws_trend.cell(row=start_row - 1, column=1, value="Errors by Type").font = BOLD_FONT
        _write_df_to_sheet(ws_trend, error_counts, start_row=start_row)

    _auto_column_width(ws_trend)

    wb.save(output_path)
    return output_path


# =========================================================================
#  Pivot Template
# =========================================================================

def generate_pivot_template(output_path: str | None = None) -> str:
    """Create a reusable pivot template for pool-level monthly analysis."""
    if output_path is None:
        output_path = os.path.join(config.TEMPLATES_DIR, "pivot_template.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # ── Instructions Sheet ──
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    instructions = [
        "LOAN PAYMENT & REMITTANCE ANALYSIS — PIVOT TEMPLATE",
        "",
        "This template is pre-configured for pool-level monthly analysis.",
        "",
        "HOW TO USE:",
        "1. Paste your cleaned payment data into the 'Data' sheet starting at A2",
        "2. Columns expected: loan_id, borrower_name, payment_date, payment_month,",
        "   scheduled_payment_amount, actual_payment_amount, payment_type,",
        "   payment_status, pool_id, variance",
        "3. Refresh the pivot summaries on the 'Pool Summary' and 'Monthly Summary' sheets",
        "4. Review the pre-built conditional formatting and filters",
        "",
        "FIELD LAYOUTS:",
        "  Pool Summary:   Rows=pool_id  |  Values=SUM(actual), COUNT(loan_id), SUM(variance)",
        "  Monthly Summary: Rows=payment_month  |  Columns=payment_status  |  Values=SUM(actual)",
    ]
    for r_idx, line in enumerate(instructions, 1):
        cell = ws_instr.cell(row=r_idx, column=1, value=line)
        if r_idx == 1:
            cell.font = Font(name="Calibri", bold=True, size=14, color="4472C4")
        elif line.startswith("HOW TO USE:") or line.startswith("FIELD LAYOUTS:"):
            cell.font = BOLD_FONT
        else:
            cell.font = NORMAL_FONT

    # ── Data Sheet (with headers only) ──
    ws_data = wb.create_sheet("Data")
    headers = [
        "loan_id", "borrower_name", "payment_date", "payment_month",
        "scheduled_payment_amount", "actual_payment_amount", "payment_type",
        "payment_status", "pool_id", "variance",
    ]
    for c_idx, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    _auto_column_width(ws_data)

    # ── Pool Summary Sheet (template) ──
    ws_pool = wb.create_sheet("Pool Summary")
    pool_headers = ["Pool ID", "Total Payments", "Total Amount $", "Avg Payment $", "Total Variance $"]
    for c_idx, h in enumerate(pool_headers, 1):
        cell = ws_pool.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for r_idx, pool in enumerate(["POOL-A", "POOL-B", "POOL-C", "POOL-D", "POOL-E", "POOL-F"], 2):
        ws_pool.cell(row=r_idx, column=1, value=pool).font = BOLD_FONT
        ws_pool.cell(row=r_idx, column=1).border = THIN_BORDER
        for c in range(2, 6):
            ws_pool.cell(row=r_idx, column=c, value="—").border = THIN_BORDER
    _auto_column_width(ws_pool)

    # ── Monthly Summary Sheet (template) ──
    ws_month = wb.create_sheet("Monthly Summary")
    month_headers = ["Month", "Posted", "Pending", "Reversed", "Total Amount $", "Variance $"]
    for c_idx, h in enumerate(month_headers, 1):
        cell = ws_month.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    _auto_column_width(ws_month)

    wb.save(output_path)
    return output_path
